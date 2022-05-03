import openpyxl
import sys

#Passing bash script argument to python script
print ('Second script argument, passing $CF_ORG: ', sys.argv[1])
print ('Third script argument, passing $CF_SPACE: ', sys.argv[2])
print ('Fourth script argument, passing $YYYYMM: ', sys.argv[3])

path = "./AuditReport-"+sys.argv[3]+".xlsx"
print ('Path: ', path)
wb = openpyxl.load_workbook(path.strip())
sheetlist = wb.sheetnames

# Add filter to the worksheets

# All Events - sheet1
sheet1 = wb[sheetlist[0]]
sheet1.auto_filter.ref = sheet1.dimensions

# User Access Changes - sheet2
sheet2 = wb[sheetlist[1]]
sheet2.auto_filter.ref = sheet2.dimensions

# Route Changes - sheet3
sheet3 = wb[sheetlist[2]]
sheet3.auto_filter.ref = sheet3.dimensions

# Service Instance Events - sheet4
sheet4 = wb[sheetlist[3]]
sheet4.auto_filter.ref = sheet4.dimensions

# Service Binding Events - sheet5
sheet5 = wb[sheetlist[4]]
sheet5.auto_filter.ref = sheet5.dimensions

# Service Events - sheet6
sheet6 = wb[sheetlist[5]]
sheet6.auto_filter.ref = sheet6.dimensions

# App Events - sheet7
sheet7 = wb[sheetlist[6]]
sheet7.auto_filter.ref = sheet7.dimensions

# ssh Events - sheet8
sheet8 = wb[sheetlist[7]]
sheet8.auto_filter.ref = sheet8.dimensions

# Space Events - sheet9
sheet9 = wb[sheetlist[8]]
sheet9.auto_filter.ref = sheet9.dimensions

# Org User Changes- sheet9
sheet10 = wb[sheetlist[9]]
sheet10.auto_filter.ref = sheet10.dimensions

print("List of Sheet names: ", sheetlist)
print("Sheet 1: ", sheet1)
print("Max row before removal: ", sheet1.max_row)

# Cleanup rows in all sheets that do not match CF_ORG or CF_SPACE

# Cleanup All Events - sheet1
rownumber = 1
for row in sheet1.iter_rows(min_row=1, max_row=sheet1.max_row, min_col=1, max_col=sheet1.max_column, values_only=True):
    rownumber += 1
for i in range(sheet1.max_row, 1, -1):
    if sheet1.cell(row=i, column=2).value != sys.argv[1] or sheet1.cell(row=i, column=3).value != sys.argv[2]:
        sheet1.delete_rows(i, 1)

# Cleanup User Access Changes - sheet2
rownumber = 1
for row in sheet2.iter_rows(min_row=1, max_row=sheet2.max_row, min_col=1, max_col=sheet2.max_column, values_only=True):
    rownumber += 1
for i in range(sheet2.max_row, 1, -1):
    if sheet2.cell(row=i, column=2).value != sys.argv[1] or sheet2.cell(row=i, column=3).value != sys.argv[2]:
        sheet2.delete_rows(i, 1)

# Cleanup Route Changes - sheet3
rownumber = 1
for row in sheet3.iter_rows(min_row=1, max_row=sheet3.max_row, min_col=1, max_col=sheet3.max_column, values_only=True):
    rownumber += 1
for i in range(sheet3.max_row, 1, -1):
    if sheet3.cell(row=i, column=2).value != sys.argv[1] or sheet3.cell(row=i, column=3).value != sys.argv[2]:
        sheet3.delete_rows(i, 1)

# Cleanup Service Instance Events - sheet4
rownumber = 1
for row in sheet4.iter_rows(min_row=1, max_row=sheet4.max_row, min_col=1, max_col=sheet4.max_column, values_only=True):
    rownumber += 1
for i in range(sheet4.max_row, 1, -1):
    if sheet4.cell(row=i, column=2).value != sys.argv[1] or sheet4.cell(row=i, column=3).value != sys.argv[2]:
        sheet4.delete_rows(i, 1)

# Cleanup Service Binding Events - sheet5
rownumber = 1
for row in sheet5.iter_rows(min_row=1, max_row=sheet5.max_row, min_col=1, max_col=sheet5.max_column, values_only=True):
    rownumber += 1
for i in range(sheet5.max_row, 1, -1):
    if sheet5.cell(row=i, column=2).value != sys.argv[1] or sheet5.cell(row=i, column=3).value != sys.argv[2]:
        sheet5.delete_rows(i, 1)

# Cleanup Service Events - sheet6
rownumber = 1
for row in sheet6.iter_rows(min_row=1, max_row=sheet6.max_row, min_col=1, max_col=sheet6.max_column, values_only=True):
    rownumber += 1
for i in range(sheet6.max_row, 1, -1):
    if sheet6.cell(row=i, column=2).value != sys.argv[1] or sheet6.cell(row=i, column=3).value != sys.argv[2]:
        sheet6.delete_rows(i, 1)

# Cleanup App Events - sheet7
rownumber = 1
for row in sheet7.iter_rows(min_row=1, max_row=sheet7.max_row, min_col=1, max_col=sheet7.max_column, values_only=True):
    rownumber += 1
for i in range(sheet7.max_row, 1, -1):
    if sheet7.cell(row=i, column=2).value != sys.argv[1] or sheet7.cell(row=i, column=3).value != sys.argv[2]:
        sheet7.delete_rows(i, 1)

# Cleanup ssh Events - sheet8
rownumber = 1
for row in sheet8.iter_rows(min_row=1, max_row=sheet8.max_row, min_col=1, max_col=sheet8.max_column, values_only=True):
    rownumber += 1
for i in range(sheet8.max_row, 1, -1):
    if sheet8.cell(row=i, column=2).value != sys.argv[1] or sheet8.cell(row=i, column=3).value != sys.argv[2]:
        sheet8.delete_rows(i, 1)

# Cleanup Space Events - sheet9
rownumber = 1
for row in sheet9.iter_rows(min_row=1, max_row=sheet9.max_row, min_col=1, max_col=sheet9.max_column, values_only=True):
    rownumber += 1
for i in range(sheet9.max_row, 1, -1):
    # Org applies filter to observe if the org has created an unknown space
    if sheet9.cell(row=i, column=2).value != sys.argv[1]:
        sheet9.delete_rows(i, 1)

# Cleanup Org User Changes - sheet10
rownumber = 1
for row in sheet10.iter_rows(min_row=1, max_row=sheet10.max_row, min_col=1, max_col=sheet10.max_column, values_only=True):
    rownumber += 1
for i in range(sheet10.max_row, 1, -1):
    # Org applies filter to observe if the org has created an unknown space
    if sheet10.cell(row=i, column=2).value != sys.argv[1]:
        sheet10.delete_rows(i, 1)

print("Max row after removal: ", sheet1.max_row)

print("Creating new file: final-AuditReport-"+str(sys.argv[3])+".xlsx...")
wb.save('final-AuditReport-'+str(sys.argv[3])+'.xlsx')

print("Cleanup report completed...")
